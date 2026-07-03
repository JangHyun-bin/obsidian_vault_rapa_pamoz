#!/usr/bin/env python3
"""WBS xlsm → Obsidian-friendly Markdown 변환기.

사용법:
    python wbs_to_md.py <input.xlsm> <output_dir>

산출물:
    - index.md : WBS 인덱스 (MOC)
    - WBS-<level>-<code>-<name>.md : L1/L2 그룹 헤더
    - WBS-L3-<code>-<name>.md : L3 태스크별 노트
    - WBS-L3-july.md : 7월 내 마감 task만 모음 (임박 뷰)
    - WBS-diff-<label>.md : diff JSON → MD 테이블
"""
import sys
import json
import re
from pathlib import Path
from datetime import datetime, date
from openpyxl import load_workbook


def normalize_date(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v)
    return s.replace(" 00:00:00", "")


def extract_tasks(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Schedule"]
    tasks = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        wbs = row[3] if len(row) > 3 else None
        if not wbs:
            continue
        name = None
        for ci in [4, 5, 6]:
            if ci < len(row) and row[ci]:
                name = row[ci]
                break
        tasks[str(wbs)] = {
            "name": str(name).strip() if name else None,
            "level": row[2] if len(row) > 2 else None,
            "start": normalize_date(row[15]) if len(row) > 15 else None,
            "end": normalize_date(row[16]) if len(row) > 16 else None,
            "total_work": row[18] if len(row) > 18 else None,
            "plan_work": row[19] if len(row) > 19 else None,
            "total_dur": row[20] if len(row) > 20 else None,
            "plan_dur": row[21] if len(row) > 21 else None,
            "owner": row[27] if len(row) > 27 else None,
            "deliverable": row[28] if len(row) > 28 else None,
        }
    wb.close()
    return tasks


def slugify(name):
    if not name:
        return "untitled"
    s = re.sub(r"[^\w가-힣\s-]", "", name).strip()
    s = re.sub(r"\s+", "-", s)
    return s[:60] if s else "untitled"


def fmt_owner(owner):
    if not owner:
        return "-"
    return str(owner).replace("\n", " ").replace("&", "&amp;").replace("|", "\\|")


def fmt_md(text):
    if not text:
        return ""
    return str(text).replace("|", "\\|").replace("\n", " ").replace("\r", "")


def render_task_md(code, t, all_tasks):
    """한 task의 .md 노트 본문."""
    name = t["name"] or "(이름 없음)"
    parts = []
    parts.append(f"# {code} {name}\n")
    parts.append("## 메타\n")
    parts.append("| 필드 | 값 |")
    parts.append("|---|---|")
    parts.append(f"| WBS 코드 | {code} |")
    parts.append(f"| 레벨 | L{t['level']} |")
    parts.append(f"| 시작 | {t['start'] or '-'} |")
    parts.append(f"| 종료 | {t['end'] or '-'} |")
    parts.append(f"| 담당 | {fmt_owner(t['owner'])} |")
    parts.append(f"| 총 작업량 | {t['total_work'] or '-'} MD |")
    parts.append(f"| 계획 작업량 | {t['plan_work'] or '-'} MD |")
    parts.append(f"| 총 기간 | {t['total_dur'] or '-'} 일 |")
    parts.append(f"| 계획 기간 | {t['plan_dur'] or '-'} 일 |")
    parts.append(f"| 산출물 | {fmt_md(t['deliverable']) or '-'} |")
    parts.append("")

    # 자식 task (L1/L2의 경우)
    children = sorted([c for c, ct in all_tasks.items()
                       if c != code and c.startswith(code + ".")])
    if children:
        parts.append("## 하위 task\n")
        parts.append("| WBS | 이름 | 시작 | 종료 | 담당 |")
        parts.append("|---|---|---|---|---|")
        for c in children:
            ct = all_tasks[c]
            parts.append(f"| [[WBS-{c}\|{c}]] | {fmt_md(ct['name']) or '-'} | "
                         f"{ct['start'] or '-'} | {ct['end'] or '-'} | "
                         f"{fmt_owner(ct['owner'])} |")
        parts.append("")

    # Linear 임베드 (Dataview 쿼리)
    parts.append("## Linear 이슈\n")
    parts.append("```dataview")
    parts.append("TABLE id, status, dueDate, assignee, project")
    parts.append(f"FROM \"\"")
    parts.append(f"WHERE contains(id, \"{code}\") OR contains(title, \"{code}\")")
    parts.append("```")
    parts.append("")

    return "\n".join(parts)


def render_index_md(tasks, src_label):
    """MOC 인덱스 노트."""
    parts = []
    parts.append(f"# WBS Index ({src_label})\n")
    parts.append(f"생성: {datetime.now().strftime('%Y-%m-%d %H:%M')} | "
                 f"전체 task: {len(tasks)}\n")
    parts.append("## L1 그룹\n")
    l1 = sorted([(c, t) for c, t in tasks.items() if t["level"] == 1],
                key=lambda x: x[0])
    for code, t in l1:
        parts.append(f"- [[WBS-L1-{code}-{slugify(t['name'])}\|{code} {t['name']}]] "
                     f"({t['start']} ~ {t['end']})")
    parts.append("\n## L2 태스크\n")
    l2 = sorted([(c, t) for c, t in tasks.items() if t["level"] == 2],
                key=lambda x: x[0])
    for code, t in l2:
        parts.append(f"- [[WBS-L2-{code}-{slugify(t['name'])}\|{code} {t['name']}]] "
                     f"({t['start']} ~ {t['end']}, 담당: {fmt_owner(t['owner'])})")
    parts.append("\n## L3 (전체 목록)\n")
    l3 = sorted([(c, t) for c, t in tasks.items() if t["level"] == 3],
                key=lambda x: x[0])
    for code, t in l3:
        parts.append(f"- [[WBS-L3-{code}-{slugify(t['name'])}\|{code} {t['name']}]] "
                     f"({t['start']} ~ {t['end']}, 담당: {fmt_owner(t['owner'])}, "
                     f"산출물: {fmt_md(t['deliverable']) or '-'})")
    parts.append("")
    return "\n".join(parts)


def render_july_md(tasks, today="2026-07-03"):
    """7월 내 마감 task만 모음 (임박 뷰)."""
    today_dt = datetime.strptime(today, "%Y-%m-%d")
    upcoming = []
    for code, t in tasks.items():
        if not t.get("end") or t["level"] != 3:
            continue
        end_dt = datetime.strptime(t["end"], "%Y-%m-%d")
        if end_dt < today_dt or end_dt > datetime(2026, 7, 31):
            continue
        owner = str(t.get("owner") or "")
        if "파모즈" not in owner and "컨소" not in owner:
            continue
        days_left = (end_dt - today_dt).days
        upcoming.append((code, t, days_left))
    upcoming.sort(key=lambda x: x[2])

    parts = []
    parts.append(f"# 7월 임박 산출물 (오늘 {today}, 파모즈+컨소 L3)\n")
    parts.append(f"총 {len(upcoming)}건. D-day 오름차순.\n")
    parts.append("| D-day | WBS | 작업명 | 시작 | 종료 | 담당 | 산출물 | 작업량 |")
    parts.append("|---|---|---|---|---|---|---|---|")
    for code, t, d in upcoming:
        urg = "🚨" if d <= 3 else ("⚠️ " if d <= 7 else "  ")
        parts.append(f"| {urg} D-{d} | [[WBS-L3-{code}-{slugify(t['name'])}\|{code}]] "
                     f"| {fmt_md(t['name'])} | {t['start']} | {t['end']} | "
                     f"{fmt_owner(t['owner'])} | {fmt_md(t['deliverable']) or '-'} | "
                     f"{t['plan_work']}MD |")
    parts.append("")
    return "\n".join(parts)


def render_diff_md(diff_data):
    """diff JSON → MD 테이블."""
    parts = []
    parts.append(f"# WBS diff: {diff_data.get('old_label','OLD')} → "
                 f"{diff_data.get('new_label','NEW')}\n")
    parts.append(f"변경 task: **{diff_data.get('diff_count',0)}건**\n")
    parts.append("| WBS | 작업 | 변경 필드 | OLD → NEW |")
    parts.append("|---|---|---|---|")
    for d in diff_data.get("diffs", []):
        if d["type"] == "신규":
            parts.append(f"| {d['code']} | {fmt_md(d['name'])} | 신규 | - |")
            continue
        if d["type"] == "삭제":
            parts.append(f"| {d['code']} | {fmt_md(d['name'])} | 삭제 | (구버전) |")
            continue
        for c in d["changes"]:
            old = c["old"] if c["old"] is not None else "(없음)"
            new = c["new"] if c["new"] is not None else "(없음)"
            parts.append(f"| {d['code']} | {fmt_md(d['name'])} | "
                         f"`{c['field']}` | {fmt_md(old)} → **{fmt_md(new)}** |")
    parts.append("")
    return "\n".join(parts)


def main():
    if len(sys.argv) < 3:
        print("Usage: wbs_to_md.py <input.xlsm> <output_dir> [diff_json]")
        sys.exit(1)
    src = Path(sys.argv[1])
    out_dir = Path(sys.argv[2])
    diff_json = Path(sys.argv[3]) if len(sys.argv) > 3 else None
    src_label = src.stem
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "L1").mkdir(exist_ok=True)
    (out_dir / "L2").mkdir(exist_ok=True)
    (out_dir / "L3").mkdir(exist_ok=True)

    tasks = extract_tasks(src)
    # 1. 인덱스
    (out_dir / "index.md").write_text(render_index_md(tasks, src_label),
                                       encoding="utf-8")
    # 2. task 노트
    for code, t in tasks.items():
        if t["level"] == 1:
            path = out_dir / "L1" / f"WBS-L1-{code}-{slugify(t['name'])}.md"
        elif t["level"] == 2:
            path = out_dir / "L2" / f"WBS-L2-{code}-{slugify(t['name'])}.md"
        elif t["level"] == 3:
            path = out_dir / "L3" / f"WBS-L3-{code}-{slugify(t['name'])}.md"
        else:
            continue
        path.write_text(render_task_md(code, t, tasks), encoding="utf-8")
    # 3. 7월 임박 뷰
    (out_dir / "WBS-july.md").write_text(render_july_md(tasks), encoding="utf-8")
    # 4. diff 뷰
    if diff_json and diff_json.exists():
        d = json.loads(diff_json.read_text(encoding="utf-8"))
        (out_dir / "WBS-diff.md").write_text(render_diff_md(d), encoding="utf-8")

    print(f"OK: {out_dir}")
    print(f"  task 수: {len(tasks)}")
    print(f"  7월 내 마감: {len([c for c, t in tasks.items() if t.get('end') and '2026-07' in str(t['end']) and t['level']==3])}건")


if __name__ == "__main__":
    main()
